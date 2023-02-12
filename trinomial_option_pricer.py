from openpyxl import load_workbook
from tree import TrinomialTree as tt
from pricer import Underlying, OptionType, ExerciseType, Option
import time
import xlwings as xw


@xw.func
@xw.arg("r", doc="interest rate")
@xw.arg("vol", doc="volatility")
@xw.arg("div", doc="dividend amount")
@xw.arg("divdate", doc="dividend ex-date")
@xw.arg("startprice", doc="starting price")
@xw.arg("strike", doc="strike")
@xw.arg("pricingdate", doc="pricing date")
@xw.arg("matdate", doc="maturity date")
@xw.arg("opttype", doc="option type")
@xw.arg("exetype", doc="exercise type")
@xw.arg("nbsteps", doc="number of steps")
@xw.func(async_mode="threading")
def pricer_trino(
    r,
    vol,
    div,
    divdate,
    startprice,
    strike,
    pricingdate,
    matdate,
    opttype,
    exetype,
    nbsteps,
):
    underlying = Underlying(r, vol, div, divdate, startprice)
    opttype = OptionType.CALL if opttype.lower() == "call" else OptionType.PUT
    exetype = ExerciseType.EUROPEAN if exetype.lower() == "european" else ExerciseType.AMERICAN
    option_ = Option(
        underlying=underlying,
        strike=strike,
        pricing_date=pricingdate,
        maturity_date=matdate,
        option_type=opttype,
        exercise_type=exetype,
        steps=int(nbsteps),
    )
    tree = tt(option_, underlying)
    tree.create()
    tree.compute_price()
    return tree.root.nfv


def price_from_excel():
    wb = load_workbook(filename="trinomial_option_pricer.xlsm")
    ws = wb["Pricer"]
    r = ws["F5"].value
    vol = ws["F6"].value
    div = ws["F7"].value
    divdate = ws["F8"].value
    startprice = ws["F4"].value
    strike = ws["F12"].value
    pricingdate = ws["F3"].value
    matdate = ws["F9"].value
    opttype = ws["F10"].value
    exetype = ws["F11"].value
    nbsteps = ws["F14"].value

    underlying = Underlying(r, vol, div, divdate, startprice)
    opttype = OptionType.CALL if opttype.lower() == "call" else OptionType.PUT
    exetype = ExerciseType.EUROPEAN if exetype.lower() == "european" else ExerciseType.AMERICAN
    option_ = Option(
        underlying=underlying,
        strike=strike,
        pricing_date=pricingdate,
        maturity_date=matdate,
        option_type=opttype,
        exercise_type=exetype,
        steps=int(nbsteps),
    )
    tree = tt(option_, underlying)
    tree.create()
    tree.compute_price()
    price = tree.root.nfv
    tree.plot()
    return price


if __name__ == "__main__":
    start_time = time.time()
    price = price_from_excel()
    print("--- %s seconds ---" % (time.time() - start_time))
    print(f"Net present value of option is: {price}")
